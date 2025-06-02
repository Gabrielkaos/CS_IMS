from django.shortcuts import render, redirect, get_object_or_404
from .models import Student, Faculty, Subject, Course, Grade, Enrollment,EnrollmentSubject
from django.contrib.auth import authenticate, login, logout
from .forms import StudentForm, UploadFileForm, FacultyForm, SubjectForm, AdminCreationForm
from django.db.models import Count
from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.auth.decorators import login_required
import openpyxl
from django.db.models import Q, F, ExpressionWrapper, FloatField
from django.contrib.auth.decorators import user_passes_test

def is_superuser(user):
    return user.is_superuser

@user_passes_test(is_superuser)
def add_admin_view(request):
    if request.method == 'POST':
        form = AdminCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'New admin created successfully.')
            return redirect('add_admin')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = AdminCreationForm()
    return render(request, 'IMS_app/add_admin.html', {'form': form})


def upload_grades(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['files']
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            # Extract metadata
            subject_description = sheet['C4'].value
            course_year = sheet['C5'].value
            sem_sy = sheet['B3'].value

            course = str(course_year).split()[0]
            year = str(course_year).split()[1]

            if len(year)==2:
                section = list(year)[1]
                year = list(year)[0]
            else:
                section = year

            semester_str = str(sem_sy).split(",")[0].lower()
            sy = str(sem_sy).split(",")[1].strip()

            if semester_str.startswith("first"):
                real_sem = 1
            elif semester_str.startswith("second"):
                real_sem = 2
            else:
                real_sem = 3


            # course_obj = Course.objects.get(name=course)
            try:
                subject_obj = Subject.objects.get(
                    description = subject_description.strip()
                )
            except:
                return redirect("grade_list")

            for row in sheet.iter_rows(min_row=9, values_only=True):
                if not row[1]:
                    continue
                student_id = str(row[1]).strip()
                print(student_id,end=" ")

                try:
                    student_obj = Student.objects.get(
                        student_id=student_id
                    )
                    print(student_obj.name,end="--")
                    
                    enrollment_obj = Enrollment.objects.get(
                        student=student_obj,
                        school_year=sy,
                        semester=real_sem,
                        year_level= int(year),
                        section=section
                    )

                    print(enrollment_obj.student.name,end="---")

                    # Link subject to enrollment
                    es_object = EnrollmentSubject.objects.get(
                        enrollment=enrollment_obj,
                        subject=subject_obj
                    )
                    mid_grade = row[2]
                    final_grade = row[3]
                    print(mid_grade)
                    print(final_grade)

                    try:
                        es_object.midterm_grade = float(mid_grade)
                        es_object.final_grade = float(final_grade)
                        es_object.save()
                    except:
                        continue
                    print("\n\nSucessful!\n\n")
                except:
                    pass

            return redirect("grade_list")

    return redirect("grade_list")



def upload_excel(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['files']
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            # Extract metadata
            course_and_section = sheet['C7'].value
            section = str(course_and_section).split()[1]
            code = sheet['C8'].value
            subject_description = sheet['C9'].value
            instructor_name = sheet['C10'].value

            sem_sy = sheet['C4'].value
            semester_str = str(sem_sy).split(",")[0].lower()
            sy = str(sem_sy).split(",")[1].strip()

            if semester_str.startswith("first"):
                real_sem = 1
            elif semester_str.startswith("second"):
                real_sem = 2
            else:
                real_sem = 3

            year_level = int(section[0]) if section[0].isdigit() else 1

            course_obj, _ = Course.objects.get_or_create(name=str(course_and_section).split()[0])

            subject_obj, _ = Subject.objects.get_or_create(
                code=code,
                defaults={'description': subject_description}
            )

            faculty_obj, _ = Faculty.objects.get_or_create(name=instructor_name)
            subject_obj.instructors.add(faculty_obj)

            for row in sheet.iter_rows(min_row=12, values_only=True):
                if not row[1] or not row[2]:
                    continue
                student_id = str(row[1]).strip()
                name = row[2].strip()

                if student_id.startswith("---"):
                    continue

                student_obj, _ = Student.objects.get_or_create(
                    student_id=student_id,
                    defaults={'name': name, 'course': course_obj}
                )
                print(section)
                enrollment_obj, _ = Enrollment.objects.get_or_create(
                    student=student_obj,
                    school_year=sy,
                    semester=real_sem,
                    defaults={
                        'year_level': year_level,
                        'section': list(section)[1] if len(section)>1 else section
                    }
                )

                # Link subject to enrollment
                EnrollmentSubject.objects.get_or_create(
                    enrollment=enrollment_obj,
                    subject=subject_obj
                )

            return redirect("student_list")

    return redirect("student_list")


# Accounts
def register(request):
    # print("Hello")
    if request.method == "POST":
        # print("HGello")
        print("I am in registerView")
        username = request.POST["username"]
        password = request.POST["password"]
        try:
            password2 = request.POST["password2"]
            # role = request.POST["role"]
        except:
            password2 = password
            # role = "student"
        

        if password != password2:
            messages.error(request, "Wrong password confirmation")
            return redirect("register")

        user, created = User.objects.get_or_create(username=username)

        if created:
            user.set_password(password)
            
            user.is_superuser = True
            user.is_staff = True
            user.save()
            messages.success(request, "User created")
            return render(request, "IMS_app/login.html")

        else:
            messages.error(request, "User already exists")
            return redirect("register")

    return render(request, "IMS_app/register.html")

def loginView(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        print("I am in loginView")
        
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid credentials.')
    return render(request, 'IMS_app/login.html')


@login_required

def dashboard(request):
    # Get selected filters from GET parameters
    selected_sy = request.GET.get('sy')
    selected_sem = request.GET.get('sem')

    # Default to most recent if not provided
    if not selected_sy:
        selected_sy = Enrollment.objects.order_by('-school_year').values_list('school_year', flat=True).first()
    if not selected_sem:
        selected_sem = Enrollment.objects.order_by('-semester').values_list('semester', flat=True).first()

    try:
        selected_sem = int(selected_sem)
    except (TypeError, ValueError):
        selected_sem = 1  # default fallback

    # Get all school years (for dropdown)
    all_school_years = Enrollment.objects.values_list('school_year', flat=True).distinct().order_by('-school_year')

    # Filtered enrollments
    filtered_enrollments = Enrollment.objects.filter(school_year=selected_sy, semester=selected_sem)

    total_students = filtered_enrollments.values('student').distinct().count()
    total_subjects = Subject.objects.count()
    total_faculty = Faculty.objects.count()
    total_enrollments = filtered_enrollments.count()

    # Top subjects (within filtered enrollments)
    top_subjects = (
        Subject.objects
        .annotate(
            student_count=Count(
                'enrollmentsubject__enrollment__student',
                filter=Q(enrollmentsubject__enrollment__in=filtered_enrollments),
                distinct=True
            )
        )
        .order_by('-student_count')[:5]
    )

    # Students per year level
    year_level_data = (
        filtered_enrollments
        .values('year_level')
        .annotate(count=Count('student'))
        .order_by('year_level')
    )

    # Students per course
    course_data = (
        filtered_enrollments
        .values('student__course__name')
        .annotate(count=Count('student'))
        .order_by('-count')
    )

    return render(request, 'IMS_app/dashboard.html', {
        'total_students': total_students,
        'total_subjects': total_subjects,
        'total_faculty': total_faculty,
        'total_enrollments': total_enrollments,
        'current_sy': selected_sy,
        'current_sem': selected_sem,
        'top_subjects': top_subjects,
        'year_level_data': year_level_data,
        'course_data': course_data,
        'all_school_years': all_school_years,
        'selected_sy': selected_sy,
        'selected_sem': selected_sem,
    })


def student_info(request, pk):
    student = get_object_or_404(Student, pk=pk)
    # Filters from the GET parameters
    selected_year = request.GET.get('school_year')
    selected_sem = request.GET.get('semester')
    selected_ylvl = request.GET.get('year_level')

    enrollments = student.enrollments.all().order_by('-school_year', '-semester')

    if selected_year:
        enrollments = enrollments.filter(school_year=selected_year)
    if selected_sem:
        enrollments = enrollments.filter(semester=selected_sem)
    if selected_ylvl:
        enrollments = enrollments.filter(year_level=selected_ylvl)

    # Get all school years, semesters, and year levels to populate the dropdowns
    school_years = student.enrollments.values_list('school_year', flat=True).distinct()
    semesters = student.enrollments.values_list('semester', flat=True).distinct()
    year_levels = student.enrollments.values_list('year_level', flat=True).distinct()

    enrollment_subjects = {}
    for enrollment in enrollments:
        subjects = EnrollmentSubject.objects.filter(enrollment=enrollment).select_related('subject')
        enrollment_subjects[enrollment] = subjects

    context = {
        'student': student,
        'enrollments': enrollment_subjects,
        'school_years': school_years,
        'semesters': semesters,
        'year_levels': year_levels,
        'selected_year': selected_year,
        'selected_semester': selected_sem,
        'selected_year_level': selected_ylvl,
    }

    return render(request, 'IMS_app/student_other_info.html', context)

def faculty_info(request, pk):
    faculty = get_object_or_404(Faculty, pk=pk)
    subjects1 = faculty.subjects.all()
    print(subjects1)
    return render(request, 'IMS_app/faculty_info.html', {'faculty': faculty,"subjects":subjects1})

def subject_info(request, pk):
    course = get_object_or_404(Subject, pk=pk)
    enrolees_on_sub = EnrollmentSubject.objects.filter(subject=course)

    return render(request, 'IMS_app/subject_info.html', {'course': course,'count1':enrolees_on_sub.count()})

@login_required
def student_list(request):
    school_year = request.GET.get('school_year')
    semester = request.GET.get('semester')
    year_level = request.GET.get('year_level')
    subject_id = request.GET.get('subject')
    course_id = request.GET.get('course')
    section = request.GET.get('section')  # NEW

    # Get all students through their enrollments
    students = Student.objects.all()

    if school_year or semester or year_level or subject_id or course_id or section:
        enrollments = Enrollment.objects.all()
        
        if school_year:
            enrollments = enrollments.filter(school_year=school_year)
        if semester:
            enrollments = enrollments.filter(semester=semester)
        if year_level:
            enrollments = enrollments.filter(year_level=year_level)
        if section:
            enrollments = enrollments.filter(section=section)
        if subject_id:
            enrollments = enrollments.filter(
                enrollmentsubject__subject__id=subject_id
            )
        
        
        if course_id:
            course_obj = Course.objects.get(id=int(course_id))
            students = Student.objects.filter(enrollments__in=enrollments,course=course_obj).distinct()
        else:
            students = Student.objects.filter(enrollments__in=enrollments).distinct()
    # Get all school years for dropdown
    school_years = Enrollment.objects.values_list('school_year', flat=True).distinct()
    subjects = Subject.objects.all()
    sections = Enrollment.objects.values_list('section', flat=True).distinct()

    form = UploadFileForm()
    return render(request, 'IMS_app/student_list.html', {
        'students': students,
        'form': form,
        'semesters': [1, 2, 3],
        'year_levels': [1, 2, 3, 4],
        'school_years': school_years,
        'subjects': subjects,
        'selected_year': school_year,
        'selected_semester': semester,
        'selected_year_level': year_level,
        'selected_subject': subject_id,
        'student_count': students.count(),
        "courses":Course.objects.all(),
        "selected_course":course_id,
        "selected_section":section,
        "sections":sections
    })


@login_required
def grade_list(request):
    school_year = request.GET.get('school_year')
    semester = request.GET.get('semester')
    year_level = request.GET.get('year_level')
    subject_id = request.GET.get('subject')
    course_id = request.GET.get('course')
    section = request.GET.get('section') 
    grade_filter = request.GET.get('grade_filter')
    print("Grade filter",grade_filter)

    # Get all students through their enrollments
    students = Student.objects.all()

    if school_year or semester or year_level or subject_id or course_id or section or grade_filter:
        enrollments = Enrollment.objects.all()
        
        if school_year:
            enrollments = enrollments.filter(school_year=school_year)
        if semester:
            enrollments = enrollments.filter(semester=semester)
        if year_level:
            enrollments = enrollments.filter(year_level=year_level)
        if section:
            enrollments = enrollments.filter(section=section)
        if subject_id:
            enrollments = enrollments.filter(
                enrollmentsubject__subject__id=subject_id
            )
        enrollment_subjects = EnrollmentSubject.objects.filter(enrollment__in=enrollments)
        filtered_student_ids = set()
        
        for es in enrollment_subjects.select_related('enrollment__student'):
            
            if grade_filter == 'gt3':
                if es.midterm_grade is not None and es.final_grade is not None:
                    avg = (es.midterm_grade + es.final_grade) / 2
                    if avg > 3.0:
                        # print(">3.0")
                        filtered_student_ids.add(es.enrollment.student.student_id)
            elif grade_filter == 'le3':
                if es.midterm_grade is not None and es.final_grade is not None:
                    avg = (es.midterm_grade + es.final_grade) / 2
                    if avg <= 3.0:
                        # print("<=3.0")
                        filtered_student_ids.add(es.enrollment.student.student_id)
            elif grade_filter == 'none':
                if es.midterm_grade is None or es.final_grade is None:
                    # print("none")
                    filtered_student_ids.add(es.enrollment.student.student_id)
            else:
                # No grade filter, include all students
                # print("super none")
                filtered_student_ids.add(es.enrollment.student.student_id)

        # If no grade filter, include all students from the filtered enrollments
        # if not grade_filter:
        #     filtered_student_ids = set(enrollments.values_list('student_id', flat=True))
        # Get the final student queryset
        # print(filtered_student_ids)
        students = Student.objects.filter(student_id__in=filtered_student_ids)
        
        if course_id:
            course_obj = Course.objects.get(id=int(course_id))
            students = students.filter(enrollments__in=enrollments,course=course_obj).distinct()
        else:
            students = students.filter(enrollments__in=enrollments).distinct()

    # Get all school years for dropdown
    school_years = Enrollment.objects.values_list('school_year', flat=True).distinct()
    subjects = Subject.objects.all()
    sections = Enrollment.objects.values_list('section', flat=True).distinct()


    form = UploadFileForm()
    return render(request, 'IMS_app/grade_list.html', {
        'students': students,
        'form': form,
        'semesters': [1, 2, 3],
        'year_levels': [1, 2, 3, 4],
        'school_years': school_years,
        'subjects': subjects,
        'selected_year': school_year,
        'selected_semester': semester,
        'selected_year_level': year_level,
        'selected_subject': subject_id,
        'student_count': students.count(),
        "courses":Course.objects.all(),
        "selected_course":course_id,
        "selected_section":section,
        "sections":sections,
        'selected_grade_filter': grade_filter
    })

def save_grades(request, pk):
    if request.method == 'POST':
        student = get_object_or_404(Student, pk=pk)
        es_objs = EnrollmentSubject.objects.filter(enrollment__student=student)

        for es in es_objs:
            midterm_key = f"midterm_{es.pk}"
            final_key = f"final_{es.pk}"
            if midterm_key in request.POST and final_key in request.POST:
                try:
                    midterm = float(request.POST[midterm_key])
                    final = float(request.POST[final_key])
                except ValueError:
                    continue

                es.midterm_grade = midterm
                es.final_grade = final
                es.save()

    return redirect('update_grades', pk=pk)

def update_grades(request, pk):
    student = get_object_or_404(Student, pk=pk)

    # Get filter params from GET
    school_year = request.GET.get('school_year')
    semester = request.GET.get('semester')
    subject_id = request.GET.get('subject')

    # For filter options
    school_years = Enrollment.objects.values_list('school_year', flat=True).distinct()
    subjects = Subject.objects.all()
    semesters = [1, 2, 3]

    # Start with all enrollment subjects related to the student
    es_objs = EnrollmentSubject.objects.filter(enrollment__student=student)

    # Apply filters if specified
    if school_year:
        es_objs = es_objs.filter(enrollment__school_year=school_year)

    if semester:
        es_objs = es_objs.filter(enrollment__semester=int(semester))

    if subject_id:
        es_objs = es_objs.filter(subject__id=int(subject_id))

    return render(request, 'IMS_app/grades.html', {
        'student': student,
        'es_objs': es_objs,  # All matching enrollment-subjects
        'selected_year': school_year,
        'selected_semester': semester,
        'selected_subject': subject_id,
        'school_years': school_years,
        'subjects': subjects,
        'semesters': semesters,
    })


@login_required
def faculty_list(request):
    faculties = Faculty.objects.all()
    form = UploadFileForm()
    return render(request, 'IMS_app/faculty_list.html', {'faculties': faculties, 'form':form})

@login_required
def subject_list(request):
    coursies = Subject.objects.all()
    return render(request, 'IMS_app/subject_list.html', {'coursies': coursies})

@login_required
def subject_create(request):
    instructors = Faculty.objects.all()
    if request.method == 'POST':
        form = SubjectForm(request.POST)
        if form.is_valid():
            # print('create form cleaned:',form.cleaned_data)
            form1 = form.save(commit=False)
            instructor = Faculty.objects.get(id=request.POST["faculties"])
            print(request.POST["faculties"])
            form1.instructor = instructor
            # print("This is the form",form)
            form1.save()
            return redirect('subject_list')
    else:
        form = SubjectForm()
    return render(request, 'IMS_app/subject_form.html', {'form': form, 'instructors':instructors})

@login_required
def subject_update(request, pk):
    course = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        form = SubjectForm(request.POST, instance=course)
        if form.is_valid():
            # print('update form cleaned:',form.cleaned_data)
            form1 = form.save(commit=False)
            # print("This is the form",form)
            form1.save()
            return redirect('subject_list')
    else:
        form = SubjectForm(instance=course)
    return render(request, 'IMS_app/subject_form.html', {'form': form})

@login_required
def student_update(request, pk):
    student = get_object_or_404(Student, pk=pk)
    coursies = Course.objects.all()
    
    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            print("Form valid student update")
            form1 = form.save()
            
            course = Course.objects.get(id=request.POST["course"])
            form1.course = course

            form1.save()
            return redirect('student_list')
    else:
        form = StudentForm(instance=student)
    return render(request, 'IMS_app/student_form.html', {'form': form,'coursies':coursies, "student":student})

@login_required
def faculty_update(request, pk):
    faculty = get_object_or_404(Faculty, pk=pk)
    if request.method == 'POST':
        form = FacultyForm(request.POST, instance=faculty)
        if form.is_valid():
            form.save()
            return redirect('faculty_list')
    else:
        form = FacultyForm(instance=faculty)
    return render(request, 'IMS_app/faculty_form.html', {'form': form})

@login_required
def student_delete(request, pk):
    student = get_object_or_404(Student, pk=pk)
    student.delete()
    return redirect('student_list')

def subject_delete(request, pk):
    course = get_object_or_404(Subject, pk=pk)
    course.delete()
    return redirect('subject_list')

def logoutView(req):

    logout(req)

    return redirect("login")


